#Excel can't be open to run this script

from openpyxl   import Workbook, load_workbook

wb = load_workbook('Test_Workbook.xlsx')

ws = wb.active 
#If you want to use a different sheet, use below for example:
#ws = wb['Sheet1']

#Prints out the value of the active sheet, in this case 'A2' 
print(ws['A13'].value)

#Changes value of a cell, in this case 'A18'
ws['A18'] = "Test"

#Creates a sheet called test
wb.create_sheet("Test")

#Prints out sheetnames
print(wb.sheetnames)

#Merges and unmerges cells 
#ws.merge_cells("A1:D1")
#ws.unmerge_cells("A1:D1")

#Inserting and deleting rows after row 7
#ws.insert_rows(7)
#ws.insert_rows(7)
#ws.delete_rows(7)
#ws.delete_rows(7)

#Inserting and deleting columns
#ws.insert_cols(2)
#ws.delete_cols(2)

#Moving rows. Format is "StartingValue:EndingValue", rows=x, cols=y, ex. rows=2 means 2 right cols=2 means 2 up 
#ws.move_range("C1:D11", rows=2, cols=2)

#Save the workbook
wb.save('Test_Workbook.xlsx')
