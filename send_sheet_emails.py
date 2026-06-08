"""
send_sheet_emails.py
--------------------
PURPOSE:
    Reads a 3-row x 3-column table from each of 5 sheets in an Excel file
    and sends a unique, personalised HTML email to each recipient via the
    Outlook desktop application.
 
HOW IT WORKS (high level):
    1. Open the Excel workbook with openpyxl.
    2. For each sheet → recipient pair defined in RECIPIENTS (below):
         a. Read the 3x3 table from that sheet.
         b. Convert the table into a styled HTML <table>.
         c. Embed the HTML table inside a full email body.
         d. Create a new Outlook mail item, fill in the fields, and send it.
 
DEPENDENCIES:
    - openpyxl   : reads Excel files         → pip install openpyxl
    - pywin32    : talks to Outlook via COM  → pip install pywin32
 
REQUIREMENTS:
    - Microsoft Outlook desktop app must be installed, open, and logged in.
    - The Excel file must exist at the path specified in EXCEL_FILE.
    - Sheet names in RECIPIENTS must exactly match the tab names in Excel
      (they are case-sensitive).
"""

import win32com.client as client
import openpyxl

# ─────────────────────────────────────────────
# CONFIG — edit these before running
# ─────────────────────────────────────────────

EXCEL_FILE = r""   # Full path to your Excel file

EMAIL_SENDER = ""          # Your email address (shown in From)

EMAIL_SUBJECT = "Your Personalised Report"  # Subject line for all emails

# Map each sheet name → recipient email.
# Order must match your sheet order (Sheet 1 → Person A, etc.)
RECIPIENTS = {
    "Sheet1": "",
    "Sheet2": "",
    "Sheet3": "",
    "Sheet4": "",
    "Sheet5": "",
}

# Rows/columns to read from each sheet (1-based, inclusive)
TABLE_START_ROW = 1
TABLE_END_ROW   = 3
TABLE_START_COL = 1
TABLE_END_COL   = 3

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def read_table(sheet, start_row, end_row, start_col, end_col):
    """Return a 2-D list of cell values from the given range."""
    table = []
    for row in sheet.iter_rows(
        min_row=start_row, max_row=end_row,
        min_col=start_col, max_col=end_col,
        values_only=True
    ):
        table.append(list(row))
    return table


def table_to_html(table):
    """Convert a 2-D list to a styled HTML table string."""
    style_table = (
        "border-collapse: collapse; font-family: Arial, sans-serif; "
        "font-size: 14px; margin: 20px 0;"
    )
    style_th = (
        "background-color: #2E4057; color: #ffffff; padding: 10px 16px; "
        "border: 1px solid #ccc; text-align: left;"
    )
    style_td = (
        "padding: 9px 16px; border: 1px solid #ccc; "
        "background-color: #f9f9f9; color: #333;"
    )
    style_td_alt = (
        "padding: 9px 16px; border: 1px solid #ccc; "
        "background-color: #eef2f7; color: #333;"
    )

    rows_html = ""
    for i, row in enumerate(table):
        cells = ""
        for cell in row:
            value = "" if cell is None else str(cell)
            if i == 0:                          # treat first row as header
                cells += f'<th style="{style_th}">{value}</th>'
            else:
                td_style = style_td if i % 2 == 1 else style_td_alt
                cells += f'<td style="{td_style}">{value}</td>'
        tag = "tr" 
        rows_html += f"<{tag}>{cells}</{tag}>\n"

    return f'<table style="{style_table}">\n{rows_html}</table>'


def build_email_body(recipient_email, sheet_name, html_table):
    """Wrap the HTML table in a full email body."""
    return f"""
    <html><body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
      <p>Hi,</p>
      <p>Please find your personalised data table below:</p>
      {html_table}
      <p style="color: #888; font-size: 12px;">
        This email was generated automatically. Data source: {sheet_name}.
      </p>
    </body></html>
    """


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print(f"Opening workbook: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    outlook = client.Dispatch("Outlook.Application")
    sent_count = 0

    for sheet_name, recipient in RECIPIENTS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [SKIP] Sheet '{sheet_name}' not found in workbook.")
            continue

        sheet = wb[sheet_name]
        table = read_table(
            sheet,
            TABLE_START_ROW, TABLE_END_ROW,
            TABLE_START_COL, TABLE_END_COL
        )

        if not any(any(cell is not None for cell in row) for row in table):
            print(f"  [SKIP] Sheet '{sheet_name}' has no data in the specified range.")
            continue

        html_table = table_to_html(table)
        body       = build_email_body(recipient, sheet_name, html_table)

        msg = outlook.CreateItem(0)          # 0 = olMailItem
        msg.To          = recipient
        msg.Subject     = EMAIL_SUBJECT
        msg.HTMLBody    = body               # Use HTMLBody, not Body
        msg.Save()                           # Save to Drafts first (avoids freeze)
        msg.Send()

        print(f"  [SENT] {sheet_name} → {recipient}")
        sent_count += 1

    wb.close()
    print(f"\nDone. {sent_count} email(s) sent.")


if __name__ == "__main__":
    main()
